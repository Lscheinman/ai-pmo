# backend/services/exporter.py
from __future__ import annotations
from typing import Iterable, Optional, Literal, List, Tuple, Dict, Any
from io import StringIO, BytesIO
import csv, zipfile, datetime as dt
from sqlalchemy.orm import Session

from db import models

try:
    import openpyxl  # optional
except Exception:
    openpyxl = None

Entity = Literal["projects", "tasks", "people", "groups", "all"]
Format = Literal["csv", "xlsx", "planner"]

# ---------- helpers ----------

def _iso(d) -> str:
    if not d:
        return ""
    return d if isinstance(d, str) else d.isoformat()

def _csv_bytes(rows: Iterable[dict], fieldnames: List[str]) -> bytes:
    sio = StringIO()
    w = csv.DictWriter(sio, fieldnames=fieldnames, extrasaction="ignore")
    w.writeheader()
    for r in rows:
        w.writerow(r)
    return sio.getvalue().encode("utf-8-sig")

def _xlsx_bytes(sheets: Dict[str, List[dict]]) -> bytes:
    if not openpyxl:
        raise RuntimeError("openpyxl not installed")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for title, rows in sheets.items():
        ws = wb.create_sheet(title=title[:31] or "Sheet")
        if not rows:
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        # simple widths
        for ci, h in enumerate(headers, 1):
            width = max(len(str(h)), *(len(str(r.get(h, ""))) for r in rows))
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(60, max(10, width + 2))
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

def _join_map(ids: List[int], m: Dict[int, str]) -> str:
    return ";".join(v for i in ids if (v := m.get(i)))

def _extract_person_ids(assignees_or_members: Any) -> List[int]:
    """
    Accepts a relationship collection where each item is either a Person or an
    association row with .person_id. Returns list of person ids (ints).
    """
    out = []
    for a in (assignees_or_members or []):
        if hasattr(a, "person_id") and a.person_id is not None:
            out.append(int(a.person_id))
        elif hasattr(a, "id") and isinstance(a.id, int):
            out.append(int(a.id))
    return out

# ---------- row builders (resolved) ----------

def _project_rows(db: Session, ids: Optional[List[int]],
                  email_by_person_id: Dict[int, str],
                  tag_label_by_id: Dict[int, str]):
    q = db.query(models.Project)
    if ids:
        q = q.filter(models.Project.id.in_(ids))
    for p in q.all():
        lead_ids = [pl.person_id for pl in getattr(p, "project_leads", []) if getattr(pl, "person_id", None)]
        tag_ids = [t.id for t in getattr(p, "tags", []) if getattr(t, "id", None)]
        yield {
            "id": p.id,
            "name": p.name or "",
            "description": p.description or "",
            "status": p.status or "",
            "start_date": _iso(p.start_date),
            "end_date": _iso(p.end_date),
            # resolved
            "lead_emails": _join_map([int(x) for x in lead_ids], email_by_person_id),
            "tag_labels": _join_map([int(x) for x in tag_ids], tag_label_by_id),
        }

def _task_rows(db: Session, ids: Optional[List[int]]):
    q = db.query(models.Task)
    if ids:
        q = q.filter(models.Task.id.in_(ids))
    for t in q.all():
        tas = getattr(t, "task_assignees", []) or []
        assignee_ids = [ta.person_id for ta in tas if getattr(ta, "person_id", None)]
        assignee_emails = ";".join(
            ta.person.email
            for ta in tas
            if getattr(ta, "person", None) and getattr(ta.person, "email", "")
        )
        assignee_names = ";".join(
            ta.person.name
            for ta in tas
            if getattr(ta, "person", None) and getattr(ta.person, "name", "")
        )

        yield {
            "id": t.id,
            "project_id": t.project_id,
            "name": t.name or "",
            "description": t.description or "",
            "status": t.status or "",
            "priority": t.priority or "",
            "start": _iso(t.start),
            "end": _iso(t.end),
            # IDs for round-tripping / joins
            "assignee_ids": ";".join(str(pid) for pid in assignee_ids),
            # Human-friendly columns
            "assignee_emails": assignee_emails,
            "assignee_names": assignee_names,
            "tag_ids": ";".join(str(tag.id) for tag in getattr(t, "tags", [])),
        }

def _person_rows(db: Session, ids: Optional[List[int]],
                 tag_label_by_id: Dict[int, str]):
    q = db.query(models.Person)
    if ids:
        q = q.filter(models.Person.id.in_(ids))
    for p in q.all():
        tag_ids = [tag.id for tag in getattr(p, "tags", []) if getattr(tag, "id", None)]
        yield {
            "id": p.id,
            "name": p.name or "",
            "email": p.email or "",
            "notes": p.notes or "",
            # resolved
            "tag_labels": _join_map([int(x) for x in tag_ids], tag_label_by_id),
        }

def _group_rows(db: Session, ids: Optional[List[int]],
                email_by_person_id: Dict[int, str],
                tag_label_by_id: Dict[int, str]):
    q = db.query(models.Group)
    if ids:
        q = q.filter(models.Group.id.in_(ids))
    for g in q.all():
        member_ids = _extract_person_ids(getattr(g, "members", []))
        tag_ids = [tag.id for tag in getattr(g, "tags", []) if getattr(tag, "id", None)]
        yield {
            "id": g.id,
            "name": g.name or "",
            "description": getattr(g, "description", "") or "",
            # resolved
            "member_emails": _join_map(member_ids, email_by_person_id),
            "tag_labels": _join_map([int(x) for x in tag_ids], tag_label_by_id),
        }

# ---------- planner mapping (tasks) ----------

def _join_emails(db: Session, assignees) -> str:
    if not assignees:
        return ""
    ids = _extract_person_ids(assignees)
    if not ids:
        return ""
    people = db.query(models.Person).filter(models.Person.id.in_(ids)).all()
    return "; ".join(p.email for p in people if p.email)

def _planner_task_rows(db: Session, ids: Optional[List[int]]):
    q = db.query(models.Task)
    if ids:
        q = q.filter(models.Task.id.in_(ids))
    for t in q.all():
        status = (t.status or "").strip().lower()
        progress = {
            "not started": "Not started",
            "in progress": "In progress",
            "blocked": "In progress",
            "complete": "Completed",
            "done": "Completed",
        }.get(status, "Not started")
        priority = {
            "low": "Low",
            "medium": "Medium",
            "high": "Important",
        }.get((t.priority or "").strip().lower(), "Medium")

        # Assigned To → emails from task_assignees
        tas = getattr(t, "task_assignees", []) or []
        assigned_to = "; ".join(
            ta.person.email
            for ta in tas
            if getattr(ta, "person", None) and getattr(ta.person, "email", "")
        )

        yield {
            "Title": t.name or f"Task {t.id}",
            "Bucket Name": "",
            "Progress": progress,
            "Priority": priority,
            "Start Date": _iso(t.start),
            "Due Date": _iso(t.end),
            "Assigned To": assigned_to,
            "Description": t.description or "",
            "Project Id": t.project_id,
            "Task Id": t.id,
        }

# ---------- façade ----------

def build_export(entity: Entity, ids: Optional[List[int]], fmt: Format, db: Session) -> Tuple[bytes, str, str]:
    """
    Returns (content_bytes, mime_type, filename)
    """
    ts = dt.datetime.now().strftime("%Y%m%d-%H%M%S")

    # Prefetch lookups once per request
    people = db.query(models.Person).all()
    email_by_person_id: Dict[int, str] = {
        int(p.id): (p.email or "").strip() for p in people if getattr(p, "id", None)
    }

    # tags may be Tag model (adjust attr name if "name" is "label" in your model)
    tags = db.query(models.Tag).all() if hasattr(models, "Tag") else []
    tag_label_by_id: Dict[int, str] = {
        int(t.id): (getattr(t, "name", None) or getattr(t, "label", "") or "").strip()
        for t in tags if getattr(t, "id", None)
    }

    projects = db.query(models.Project).all()
    project_name_by_id: Dict[int, str] = {
        int(p.id): (p.name or "").strip() for p in projects if getattr(p, "id", None)
    }

    if fmt == "planner":
        if entity not in ("tasks", "all"):
            raise ValueError("Planner format is only supported for tasks or all.")
        rows = list(_planner_task_rows(db, ids if entity == "tasks" else None))
        fields = ["Title","Bucket Name","Progress","Priority","Start Date","Due Date","Assigned To","Description","Project Id","Task Id"]
        data = _csv_bytes(rows, fields)
        return data, "text/csv", f"tasks-planner-{ts}.csv"

    if fmt == "csv":
        if entity == "projects":
            rows = list(_project_rows(db, ids, email_by_person_id, tag_label_by_id))
            fields = rows[0].keys() if rows else [
                "id","name","description","status","start_date","end_date","lead_emails","tag_labels"
            ]
            return _csv_bytes(rows, list(fields)), "text/csv", f"projects-{ts}.csv"

        if entity == "tasks":
            rows = list(_task_rows(db, ids))
            fields = rows[0].keys() if rows else [
                "id","project_id","project_name","name","description","status","priority","start","end","assignee_emails","tag_labels"
            ]
            return _csv_bytes(rows, list(fields)), "text/csv", f"tasks-{ts}.csv"

        if entity == "people":
            rows = list(_person_rows(db, ids, tag_label_by_id))
            fields = rows[0].keys() if rows else ["id","name","email","notes","tag_labels"]
            return _csv_bytes(rows, list(fields)), "text/csv", f"people-{ts}.csv"

        if entity == "groups":
            rows = list(_group_rows(db, ids, email_by_person_id, tag_label_by_id))
            fields = rows[0].keys() if rows else ["id","name","description","member_emails","tag_labels"]
            return _csv_bytes(rows, list(fields)), "text/csv", f"groups-{ts}.csv"

        if entity == "all":
            bio = BytesIO()
            with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as z:
                z.writestr(
                    f"projects-{ts}.csv",
                    _csv_bytes(
                        list(_project_rows(db, None, email_by_person_id, tag_label_by_id)),
                        ["id","name","description","status","start_date","end_date","lead_emails","tag_labels"],
                    ),
                )
                z.writestr(
                    f"tasks-{ts}.csv",
                    _csv_bytes(
                        list(_task_rows(db, ids)),
                        ["id","project_id","project_name","name","description","status","priority","start","end","assignee_emails","tag_labels"],
                    ),
                )
                z.writestr(
                    f"people-{ts}.csv",
                    _csv_bytes(
                        list(_person_rows(db, None, tag_label_by_id)),
                        ["id","name","email","notes","tag_labels"],
                    ),
                )
                z.writestr(
                    f"groups-{ts}.csv",
                    _csv_bytes(
                        list(_group_rows(db, None, email_by_person_id, tag_label_by_id)),
                        ["id","name","description","member_emails","tag_labels"],
                    ),
                )
            return bio.getvalue(), "application/zip", f"export-{ts}.zip"

    if fmt == "xlsx":
        sheets: Dict[str, List[dict]] = {}
        if entity in ("projects", "all"):
            sheets["Projects"] = list(_project_rows(db, ids if entity=="projects" else None, email_by_person_id, tag_label_by_id))
        if entity in ("tasks", "all"):
            sheets["Tasks"] = list(_task_rows(db, ids if entity=="tasks" else None))
        if entity in ("people", "all"):
            sheets["People"] = list(_person_rows(db, ids if entity=="people" else None, tag_label_by_id))
        if entity in ("groups", "all"):
            sheets["Groups"] = list(_group_rows(db, ids if entity=="groups" else None, email_by_person_id, tag_label_by_id))
        data = _xlsx_bytes(sheets)
        return data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"export-{ts}.xlsx"

    raise ValueError("Unsupported format")
